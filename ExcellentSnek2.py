#! /usr/bin/python3

import string

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

# import your file 

wb = load_workbook(filename = 'IMPORTDATA.xlsx')

# Assuming your sheet is named Sheet1 

sheet_ranges = wb['Sheet1']
sheet = wb.active

# count num rows
row_count = sheet.max_row
# count num columns
col_count = sheet.max_column

# print a count to aid in debugging

print("Total Rows are",row_count)
print("Total Columns are",col_count)

# variables to start the loop
# startrow will read from second column to allow for a header
# startcolumn is just first column

startrow = 2
startcolumn = 1

while row_count >= startrow:

    sheet.cell(row=1, column=1).value = sheet.cell(row=1, column=1).value
    
    col2 = startcolumn + 1
    col3 = startcolumn + 2
    col4 = startcolumn + 3

# This version will eliminate extra bullets, unless all cells are empty


    if str(sheet.cell(row=startrow, column=col2).value) == 'None':
        data2 = '<UL>\n'
    else:
        data2 = str(sheet.cell(row=startrow, column=col2).value)
        data2 = '<UL>\n<LI>\n' + data2 + '\n</LI>\n'
        
        
    if str(sheet.cell(row=startrow, column=col3).value) == 'None':
        data3 = ""
    else:
        data3 = str(sheet.cell(row=startrow, column=col3).value) 
        data3 = '\n<LI>\n' + data3 + '\n</LI>\n' 
        
    if str(sheet.cell(row=startrow, column=col4).value) == 'None':
        data4 = "\n</UL>\n\n"
    else:
        data4 = str(sheet.cell(row=startrow, column=col4).value) 
        data4 = '\n<LI>\n' + data4 + '\n</LI>\n</UL>\n\n' 
 
       
       
# Now we combine the data sets to create one cell all html will work even if no data in this format
    
    combined = data2 + data3 + data4

    sheet.cell(row=startrow, column=2).value = combined
    
    startrow = startrow + 1
    
sheet.cell(row=1, column=1).value = "IDENT"
sheet.cell(row=1, column=2).value = "COMBINED HTML DATA"


# clean extra cells
startrow = 1
startcolumn = 1

while row_count >= startrow:

    
    col3 = startcolumn + 2
    col4 = startcolumn + 3
    col5 = startcolumn + 4


    sheet.cell(row=startrow, column=col3).value = None
    sheet.cell(row=startrow, column=col4).value = None
    sheet.cell(row=startrow, column=col5).value = None


    startrow = startrow + 1   



# save new file called EXPORTFILE.xlsx

dest_filename = 'EXPORTFILE.xlsx'
wb.save(filename = dest_filename)

# let us know it's done

print ("work complete")        
